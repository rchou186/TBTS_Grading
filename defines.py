###############################################################################
### Lookup for the model related battery information                        ###
### Define filled colors for excel cells                                    ###
###############################################################################

from openpyxl.styles import Font, PatternFill, Color, Border, Side

def get_max_module(battery_model):    #for get the MaxModule, it's a dictionary for lookup from input BatMod
    return {1:34 ,2:28 ,3:34, 4:32, 5:40, 6:40, 7:40, 8:20}.get(battery_model,0)

def get_model_info(battery_model):    #for put in the Table
    return {1:'34M',2:'28M',3:'18M+16M', 4:'32M', 5:'40M', 6:'34M+6M', 7:'20M+20M', 8:'20M'}.get(battery_model,'error')

def get_model_suffix(battery_model):  #for put at the suffix of output excel filename
    return {1:'-(34)',2:'-(28)',3:'-(18+16)', 4:'-(32)', 5:'-(40)', 6:'-(34+6)', 7:'-(20+20)', 8:'-(20)'}.get(battery_model, '-()')

def get_model_result(battery_model):  #for put in the Result page
    return {1:'[34]',2:'[28]',3:'[18]+[16]', 4:'[32]', 5:'[40]', 6:'[34]+[6]', 7:'[20]+[20]', 8:'[20]'}.get(battery_model, '[]')

# define colors in BIN cells
GreenFill = PatternFill(fill_type='solid', fgColor='66FF66')    # color for G[1], K or D
PurpleFill = PatternFill(fill_type='solid', fgColor='CC99FF')   # color for G[2], P or F
OrangeFill = PatternFill(fill_type='solid', fgColor='FF9900')   # color for G[3], O or G
WhiteFill = PatternFill(fill_type='solid', fgColor='C0C0C0')    # color for G[4], W or N
BlueFill = PatternFill(fill_type='solid', fgColor='3366FF')     # color for G[5], E or Y
RedFill = PatternFill(fill_type='solid', fgColor='FF0000')      # color for G[6], H or Z
LGreenFill = PatternFill(fill_type='solid', fgColor='8AE600')   # color for G[1] without VDT check
LBlueFill = PatternFill(fill_type='solid', fgColor='7F7FFF')    # color for G[5] without full charge
LRedFill = PatternFill(fill_type='solid', fgColor='FF0080')     # color for G[6] without full charge
TitleFill = PatternFill(fill_type='solid', fgColor='DCE6F1')    # title of result table
NGFill = PatternFill(fill_type='solid', fgColor='F2F2F2')       # background of result table
OKFill = PatternFill(fill_type='solid', fgColor='A6A6A6')       # good module of result table
BlankFill = PatternFill()                                       # a null fill

# define sign constant lists
sign1 = [' ', '-', '+', '+', '+', '+', '+', '-', '-']           #sign at M01 position
sign2 = [' ', '+', '-', '-', '-', '-', '-', '+', '+']           #sign at M-last position
