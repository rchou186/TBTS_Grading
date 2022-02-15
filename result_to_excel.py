###############################################################################
### Make result table sheet and result for print sheet in excel             ###
### Module to install: pandas, openpyxl                                     ###
###############################################################################

from openpyxl import Workbook
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.table import TableStyleElement
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from defines import *

VDELTAMAX_THRES = 2.4


def result_to_excel(table, ColorList, battery_model, version, result_filename):

    max_module = len(table)

    wb = Workbook()
    ws_table = wb.active
    ws_table.title = "Table"
    ft_Arial_12 = Font(name='Arial', size=12)
    ft_Arial_8 = Font(name='Arial', size=8)
    ft_Calibri_14 = Font(name='Calibri', size=14)

    # write table to worksheet
    for r in dataframe_to_rows(table, index=False, header=True):
        ws_table.append(r)

    # change the font of ws sheet by for loops of each cells
    # also change the format of date, time and percentage of the cells respectively
    for row in range(1, ws_table.max_row+1):
        for col in range(1, ws_table.max_column+1):
            ws_table.cell(row, col).font = ft_Arial_12
            if (col in [6, 9, 12]):                     # change the number_format to 0.00% for percentage
                ws_table.cell(row, col).number_format = '0.00%'
            if(col in [10, 11, 18]):                    # change the number_format to H:MM:SS for time
                ws_table.cell(row, col).number_format = 'H:MM:SS'
            if (col == 26):
                ws_table.cell(row, col).number_format = 'MM-DD-YYYY'
            if (col == 27):
                ws_table.cell(row, col).number_format = 'HH:MM:SS'
    for i in range(len(ColorList)):
        ws_table.cell(i+2, 21).fill = ColorList[i]

    # assign column width to each column
    ws_table.column_dimensions['A'].width = 30.5
    ws_table.column_dimensions['B'].width = 14.5
    ws_table.column_dimensions['D'].width = 20
    ws_table.column_dimensions['F'].width = 11
    ws_table.column_dimensions['I'].width = 11
    ws_table.column_dimensions['J'].width = 11
    ws_table.column_dimensions['K'].width = 11
    ws_table.column_dimensions['L'].width = 11
    ws_table.column_dimensions['R'].width = 11
    ws_table.column_dimensions['Z'].width = 14
    ws_table.column_dimensions['AA'].width = 11
    ws_table.column_dimensions['AG'].width = 12
    ws_table.column_dimensions['AH'].width = 12

    # generate Result worksheet

    # define borders
    top_thick_border = Border(top=Side(border_style='thick'))
    bottom_thick_border = Border(bottom=Side(border_style='thick'))
    left_thick_border = Border(left=Side(border_style='thick'))
    right_thick_border = Border(right=Side(border_style='thick'))
    topleft_thick_border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'))
    topright_thick_border = Border(top=Side(border_style='thick'), right=Side(border_style='thick'))
    bottomleft_thick_border = Border(bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    bottomright_thick_border = Border(bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    double_thin_border = Border(top=Side(border_style='double'), bottom=Side(border_style='double'),
                                left=Side(border_style='double'), right=Side(border_style='double'))

    wb.create_sheet('Result')
    ws_result = wb['Result']

    # set the column width
    ws_result.column_dimensions['A'].width = 0.5
    for col in range(2, 50):  # from B to AT
        ws_result.column_dimensions[get_column_letter(col)].width = 4

    # set the row height
    ws_result.row_dimensions[1].height = 17.25
    ws_result.row_dimensions[2].height = 17.25
    ws_result.row_dimensions[3].height = 24
    ws_result.row_dimensions[4].height = 70
    ws_result.row_dimensions[5].height = 17.25
    ws_result.row_dimensions[6].height = 17.25
    ws_result.row_dimensions[7].height = 110
    ws_result.row_dimensions[9].height = 35

    # set all the font
    for row in range(2, 10):
        for col in range(2, 50):
            ws_result.cell(row, col).font = ft_Arial_12
            if row in range(2, 4):
                ws_result.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')

    # write basic data into cells
    ws_result.cell(1, 2).value = version
    ws_result.cell(1, 2).font = Font(name='新細明體', size=12)
    ws_result.cell(1, 2).alignment = Alignment(horizontal='left', vertical='center')
    ws_result.cell(1, 6).value = str(VDELTAMAX_THRES)
    ws_result.cell(1, 6).font = Font(name='新細明體', size=12)
    ws_result.cell(1, 6).alignment = Alignment(horizontal='left', vertical='center')
    ws_result.cell(2, 2).value = table.loc[0, 'Battery']
    ws_result.cell(2, 6).value = table.loc[0, 'B_SN']
    #ws_result.cell(2, 16).value = BatteryModel
    ws_result.cell(2, 17).value = get_model_result(battery_model)
    ws_result.cell(3, 6).value = table.loc[0, 'Date']  # .strftime("%m-%d-%Y")
    ws_result.cell(3, 10).value = table.loc[0, 'Time']  # .strftime("%H:%M")
    ws_result.cell(3, 13).value = table.loc[0, 'Station']
   
    # get middle module number
    if battery_model == 3:      # 18+16
        middle_module = 18
        ws_result.column_dimensions['V'].width = 0.1
    elif battery_model == 6:    # 34+6
        middle_module = 34
        ws_result.column_dimensions['AL'].width = 0.1
    elif battery_model == 7:    # 20+20
        middle_module = 20
        ws_result.column_dimensions['X'].width = 0.1
    else:
        middle_module = max_module

    if middle_module == max_module:
        
        # one piece--------------------------------------------------------------------------------------
        for i in range(0, max_module+2):
            ws_result.cell(2, i+2).fill = TitleFill
            for j in range(3, 7):
                ws_result.cell(j, i+2).fill = NGFill
        
        # draw border
        for i in range(1, max_module+1):
            ws_result.cell(2, i+2).border = top_thick_border  # top thick border
            ws_result.cell(6, i+2).border = bottom_thick_border  # bottom thick border
        for i in range(3, 6):
            ws_result.cell(i, 2).border = left_thick_border  # left
            ws_result.cell(i, max_module+3).border = right_thick_border  # right
        for i in range(1, max_module+1):
            ws_result.cell(4, i+2).border = double_thin_border
        ws_result.cell(2, 2).border = topleft_thick_border  # topleft corner
        ws_result.cell(2, max_module+3).border = topright_thick_border  # topright corner
        ws_result.cell(6, 2).border = bottomleft_thick_border  # bottomleft corner
        ws_result.cell(6, max_module+3).border = bottomright_thick_border  # bottomright cornerelse:
        
        # write sign and bar
        ws_result.cell(3, 3).value = sign1[battery_model]
        ws_result.cell(3, 3).font = Font(name='Arial', size=16, bold=True)
        ws_result.cell(3, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws_result.cell(3, 3).fill = BlankFill
        ws_result.cell(3, max_module+2).value = sign2[battery_model]
        ws_result.cell(3, max_module+2).font = Font(name='Arial', size=16, bold=True)
        ws_result.cell(3, max_module+2).alignment = Alignment(horizontal='center', vertical='center')
        ws_result.cell(3, max_module+2).fill = BlankFill
        ws_result.cell(4, max_module+3).value = '▌\n▌\n▌'
        ws_result.cell(4, max_module+3).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # write data
        for i in range(0, max_module):
            ws_result.cell(4, i+3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if table.loc[i, 'BIN'] == 'ERR' or table.loc[i, 'BIN'] == 'SER':
                ws_result.cell(4, i+3).value = '\n\n\n\nERR'
                ws_result.cell(4, i+3).font = ft_Arial_8
            else:
                ws_result.cell(4, i+3).value = table.loc[i, 'BIN'][0:2]
                if ws_result.cell(4, i+3).value == 'NG':
                    ws_result.cell(4, i+3).font = ft_Arial_8
                else:
                    ws_result.cell(4, i+3).fill = OKFill
            ws_result.cell(5, i+3).value = '{0:02d}'.format(i+1)
            ws_result.cell(7, i+3).value = table.loc[i, 'M_SN']
            ws_result.cell(7, i+3).alignment = Alignment(textRotation=90)
            ws_result.cell(9, i+3).value = table.loc[i, 'VDeltaMax']
            ws_result.cell(9, i+3).alignment = Alignment(textRotation=90)
            if table.loc[i, 'VDeltaMax'] > VDELTAMAX_THRES:
                ws_result.cell(9, i+3).fill = OKFill
    else:
        
        # first half------------------------------------------------------------------------------------
        for i in range(0, middle_module+2):
            ws_result.cell(2, i+2).fill = TitleFill
            for j in range(3, 7):
                ws_result.cell(j, i+2).fill = NGFill
        
        # draw border
        for i in range(1, middle_module+1):
            ws_result.cell(2, i+2).border = top_thick_border  # top thick border
            ws_result.cell(6, i+2).border = bottom_thick_border  # bottom thick border
        for i in range(3, 6):
            ws_result.cell(i, 2).border = left_thick_border  # left
            ws_result.cell(i, middle_module+3).border = right_thick_border  # right
        for i in range(1, middle_module+1):
            ws_result.cell(4, i+2).border = double_thin_border
        ws_result.cell(2, 2).border = topleft_thick_border  # topleft corner
        ws_result.cell(2, middle_module+3).border = topright_thick_border  # topright corner
        ws_result.cell(6, 2).border = bottomleft_thick_border  # bottomleft corner
        ws_result.cell(6, middle_module+3).border = bottomright_thick_border  # bottomright cornerelse:
        
        # write sign and bar
        ws_result.cell(3, 3).value = sign1[battery_model]
        ws_result.cell(3, 3).font = Font(name='Arial', size=16, bold=True)
        ws_result.cell(3, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws_result.cell(3, 3).fill = BlankFill
        ws_result.cell(3, middle_module+2).value = sign2[battery_model]
        ws_result.cell(3, middle_module+2).font = Font(name='Arial', size=16, bold=True)
        ws_result.cell(3, middle_module+2).alignment = Alignment(horizontal='center', vertical='center')
        ws_result.cell(3, middle_module+2).fill = BlankFill
        ws_result.cell(4, middle_module+3).value = '▌\n▌\n▌'
        ws_result.cell(4, middle_module+3).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # write data
        for i in range(0, middle_module):
            ws_result.cell(4, i+3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if table.loc[i, 'BIN'] == 'ERR' or table.loc[i, 'BIN'] == 'SER':
                ws_result.cell(4, i+3).value = '\n\n\n\nERR'
                ws_result.cell(4, i+3).font = ft_Arial_8
            else:
                ws_result.cell(4, i+3).value = table.loc[i, 'BIN'][0:2]
                if ws_result.cell(4, i+3).value == 'NG':
                    ws_result.cell(4, i+3).font = ft_Arial_8
                else:
                    ws_result.cell(4, i+3).fill = OKFill
            ws_result.cell(5, i+3).value = '{0:02d}'.format(i+1)
            ws_result.cell(7, i+3).value = table.loc[i, 'M_SN']
            ws_result.cell(7, i+3).alignment = Alignment(textRotation=90)
            ws_result.cell(9, i+3).value = table.loc[i, 'VDeltaMax']
            ws_result.cell(9, i+3).alignment = Alignment(textRotation=90)
            if table.loc[i, 'VDeltaMax'] > VDELTAMAX_THRES:
                ws_result.cell(9, i+3).fill = OKFill
        
        # second half---------------------------------------------------------------------------------
        for i in range(middle_module+3, max_module+5):
            ws_result.cell(2, i+2).fill = TitleFill
            for j in range(3, 7):
                ws_result.cell(j, i+2).fill = NGFill
        
        # draw border
        for i in range(middle_module+4, max_module+4):
            ws_result.cell(2, i+2).border = top_thick_border  # top thick border
            ws_result.cell(6, i+2).border = bottom_thick_border  # bottom thick border
        for i in range(3, 6):
            ws_result.cell(i, middle_module+5).border = left_thick_border  # left
            ws_result.cell(i, max_module+6).border = right_thick_border  # right
        for i in range(middle_module+4, max_module+4):
            ws_result.cell(4, i+2).border = double_thin_border
        ws_result.cell(2, middle_module+5).border = topleft_thick_border  # topleft corner
        ws_result.cell(2, max_module+6).border = topright_thick_border  # topright corner
        ws_result.cell(6, middle_module+5).border = bottomleft_thick_border  # bottomleft corner
        ws_result.cell(6, max_module+6).border = bottomright_thick_border  # bottomright cornerelse:
        
        # write sign and bar
        ws_result.cell(3, middle_module+6).value = sign1[battery_model]
        ws_result.cell(3, middle_module+6).font = Font(name='Arial', size=16, bold=True)
        ws_result.cell(3, middle_module+6).alignment = Alignment(horizontal='center', vertical='center')
        ws_result.cell(3, middle_module+6).fill = BlankFill
        ws_result.cell(3, max_module+5).value = sign2[battery_model]
        ws_result.cell(3, max_module+5).font = Font(name='Arial', size=16, bold=True)
        ws_result.cell(3, max_module+5).alignment = Alignment(horizontal='center', vertical='center')
        ws_result.cell(3, max_module+5).fill = BlankFill
        ws_result.cell(4, max_module+6).value = '▌\n▌\n▌'
        ws_result.cell(4, max_module+6).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # write data
        for i in range(middle_module, max_module):
            ws_result.cell(4, i+6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if table.loc[i, 'BIN'] == 'ERR' or table.loc[i, 'BIN'] == 'SER':
                ws_result.cell(4, i+6).value = '\n\n\n\nERR'
                ws_result.cell(4, i+6).font = ft_Arial_8
            else:
                ws_result.cell(4, i+6).value = table.loc[i, 'BIN'][0:2]
                if ws_result.cell(4, i+6).value == 'NG':
                    ws_result.cell(4, i+6).font = ft_Arial_8
                else:
                    ws_result.cell(4, i+6).fill = OKFill
            ws_result.cell(5, i+6).value = '{0:02d}'.format(i+1)
            ws_result.cell(7, i+6).value = table.loc[i, 'M_SN']
            ws_result.cell(7, i+6).alignment = Alignment(textRotation=90)
            ws_result.cell(9, i+6).value = table.loc[i, 'VDeltaMax']
            ws_result.cell(9, i+6).alignment = Alignment(textRotation=90)
            if table.loc[i, 'VDeltaMax'] > VDELTAMAX_THRES:
                ws_result.cell(9, i+3).fill = OKFill

    # move the Result worksheet 5 position ahead to the first position
    wb.move_sheet('Result', offset=-5)  
    
    wb.save(result_filename + '.xlsx')
